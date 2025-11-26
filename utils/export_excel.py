"""Excel export helpers for product catalog and inventory listings."""
from __future__ import annotations

from typing import Iterable, Mapping

from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side
from openpyxl.utils import get_column_letter

HEADERS = [
    "SKU",
    "Descripción",
    "Tipo",
    "Departamento",
    "Proveedor",
    "Precio Costo",
    "Precio Venta",
    "Precio Mayoreo",
    "Inventario",
    "Min",
    "Max",
    "Fecha actualización",
]


def _row_from_product(product: Mapping[str, object]) -> list[object]:
    return [
        product.get("sku", ""),
        product.get("name", ""),
        product.get("sale_type", ""),
        product.get("department") or product.get("category") or "",
        product.get("provider") or "",
        product.get("cost", ""),
        product.get("price", ""),
        product.get("price_wholesale", ""),
        product.get("stock", ""),
        product.get("min_stock", ""),
        product.get("max_stock", ""),
        product.get("updated_at", ""),
    ]


def _auto_width(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                cell_length = len(str(cell.value)) if cell.value is not None else 0
                max_length = max(max_length, cell_length)
            except Exception:  # noqa: BLE001
                continue
        ws.column_dimensions[column].width = min(max_length + 2, 50)


def export_product_catalog_to_excel(products: Iterable[Mapping[str, object]], filepath: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    ws.append(HEADERS)
    thin = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.border = thin
    for product in products:
        ws.append(_row_from_product(product))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = thin
    _auto_width(ws)
    wb.save(filepath)


def export_inventory_to_excel(products: Iterable[Mapping[str, object]], filepath: str) -> None:
    inventory_products = [p for p in products if p.get("uses_inventory", True)]
    export_product_catalog_to_excel(inventory_products, filepath)
