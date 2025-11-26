"""Utilities to export customer catalogues to CSV or Excel."""
from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable, Mapping

from openpyxl import Workbook


EXPORT_COLUMNS = [
    "first_name",
    "last_name",
    "email",
    "email_fiscal",
    "phone",
    "domicilio1",
    "domicilio2",
    "colonia",
    "municipio",
    "estado",
    "pais",
    "codigo_postal",
    "notes",
    "credit_limit",
    "credit_balance",
    "last_payment_ts",
    "last_payment_amount",
    "rfc",
    "razon_social",
    "regimen_fiscal",
]


def _iter_rows(customers: Iterable[Mapping[str, object]]):
    for customer in customers:
        yield [customer.get(col, "") for col in EXPORT_COLUMNS]


def export_customers_to_csv(customers: Iterable[Mapping[str, object]], filepath: str | Path) -> None:
    path = Path(filepath)
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(EXPORT_COLUMNS)
        writer.writerows(_iter_rows(customers))


def export_customers_to_excel(customers: Iterable[Mapping[str, object]], filepath: str | Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(EXPORT_COLUMNS)
    for row in _iter_rows(customers):
        ws.append(row)
    wb.save(filepath)
